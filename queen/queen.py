import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

n = int(input("请输入皇后个数:(建议输入数值在11及以下)\n"))

# 递归解决N皇后问题并且把分布结果存储在res数组中
column = [False] * (n + 1)
diag = [False] * 2 * (n + 1)
diag2 = [False] * 2 * (n + 1)
res = []
pres = [0] * (n + 1)


def function(n, k):
    if k == n + 1:
        res.append(pres.copy())
        return
    for i in range(1, n + 1):
        if column[i] is False and diag[k - i + n] is False and diag2[i + k] is False:
            column[i] = True
            diag[k - i + n] = True
            diag2[i + k] = True
            pres[k] = i
            function(n, k + 1)
            column[i] = False
            diag[k - i + n] = False
            diag2[i + k] = False
    return


# 开始操作excel表格
rowHeight = 32  # 行高
columnWidth = 8  # 列宽
function(n, 1)
wb1 = op.Workbook()
ws = wb1.active
for k in range(0, len(res) - 1, 2):
    for i in range(0, n):
        for j in range(1, n + 1):
            tempCell = ws.cell(i + 1 + k * 6, j)
            tempCell2 = ws.cell(i + 1 + k * 6, j + 12)
            tempCell.font = Font(color="111111", size='25')
            tempCell2.font = Font(color="111111", size='25')
            if (i + j) % 2 == 0:
                tempCell.fill = PatternFill(fill_type='solid', start_color='111111')
                tempCell.font = Font(color="ffffff", size='25')
                tempCell2.fill = PatternFill(fill_type='solid', start_color='111111')
                tempCell2.font = Font(color="ffffff", size='25')
            if res[k][i + 1] == j:
                tempCell.alignment = Alignment(horizontal="center", vertical='center')
                tempCell.value = '\u2655'
            if res[k + 1][i + 1] == j:
                tempCell2.alignment = Alignment(horizontal="center", vertical='center')
                tempCell2.value = '\u2655'
# 设置行高列宽
for i in range(1, ws.max_row + 1):
    ws.row_dimensions[i].height = rowHeight
for i in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(i)].width = columnWidth
wb1.save("./fill.xlsx")
print("输出完成")
